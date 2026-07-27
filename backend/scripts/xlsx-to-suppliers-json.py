#!/usr/bin/env python3
"""Extrait l'onglet Base_Fournisseurs d'un export Excel de suivi fournisseurs (format du fichier
utilisé par CCG/Soguipal aujourd'hui) et produit un JSON propre consommable par
import-suppliers.js. Ne touche pas à la base — c'est le script Node qui écrit en base.

Usage : python xlsx-to-suppliers-json.py <fichier.xlsx> <sortie.json>

Le fichier source a une ligne par OFFRE (un même fournisseur peut apparaître plusieurs fois sous
le même Code_Fournisseur pour des catégories/produits différents) ; ce script fusionne ces lignes
en un seul enregistrement par fournisseur (une ligne = une fiche), cohérent avec le modèle de
données de CCG Flow.
"""
import sys
import json
import openpyxl

HEADER_MAP = {
    'Code_Fournisseur': 'code',
    'Origine': 'origine',
    'Groupe': 'categorie',
    'Fournisseur': 'nom',
    'Désignation': 'produits_offres',
    'Pays': 'pays',
    'Téléphone': 'contact_tel',
    'Email': 'contact_email',
    'Contact': 'contact_nom',
    'Mode de paiement': 'mode_paiement',
    'Delai de paiement': 'conditions_paiement',
    'Existence contrat': 'a_contrat',
    'Commentaires': 'commentaires',
}


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def normalize_origine(v):
    v = clean(v)
    if v is None:
        return None
    return 'Import' if v.lower() == 'import' else 'Local' if v.lower() == 'local' else v


def normalize_bool(v):
    v = clean(v)
    if v is None:
        return None
    return v.lower() == 'oui'


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    src, out = sys.argv[1], sys.argv[2]

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb['Base_Fournisseurs']
    headers = [c.value for c in ws[1]]
    cols = {h: i for i, h in enumerate(headers) if h in HEADER_MAP}

    merged = {}  # clé de fusion -> enregistrement
    order = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {}
        for header, idx in cols.items():
            key = HEADER_MAP[header]
            val = row[idx]
            if key == 'origine':
                val = normalize_origine(val)
            elif key == 'a_contrat':
                val = normalize_bool(val)
            elif key == 'contact_tel':
                val = clean(val)
            else:
                val = clean(val)
            rec[key] = val

        if not rec.get('nom'):
            continue  # ligne vide ou sans nom exploitable

        merge_key = rec.get('code') or f"__no_code__{rec['nom'].lower()}"
        if merge_key not in merged:
            merged[merge_key] = rec
            order.append(merge_key)
        else:
            existing = merged[merge_key]
            for multi_key in ('categorie', 'produits_offres'):
                vals = [v for v in (existing.get(multi_key), rec.get(multi_key)) if v]
                existing[multi_key] = ' / '.join(dict.fromkeys(vals)) or None
            for single_key in ('contact_tel', 'contact_email', 'contact_nom', 'mode_paiement',
                                'conditions_paiement', 'a_contrat', 'commentaires', 'pays', 'origine'):
                if not existing.get(single_key) and rec.get(single_key) is not None:
                    existing[single_key] = rec[single_key]

    records = [merged[k] for k in order]
    records.sort(key=lambda r: r['nom'].lower())

    with open(out, 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    print(f"{len(records)} fournisseurs extraits -> {out}")


if __name__ == '__main__':
    main()
