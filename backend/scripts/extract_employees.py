"""Extrait et nettoie les employés depuis le fichier Excel RH source.
Usage : python extract_employees.py "<chemin xlsx>" employees_import.json
Ne touche pas à la base de données — produit juste un JSON intermédiaire propre,
lu ensuite par scripts/import-employees.js.
"""
import json
import re
import sys
import unicodedata
from datetime import datetime, date

import openpyxl


def split_nom_complet(full_name):
    full_name = (full_name or '').strip()
    if not full_name:
        return '', ''
    tokens = full_name.split()
    if len(tokens) == 1:
        return '', tokens[0]

    def is_upper_token(t):
        letters = [c for c in t if c.isalpha()]
        return bool(letters) and all(c.upper() == c for c in letters)

    # Cherche la suite de tokens en MAJUSCULES en fin de chaîne (le nom de famille).
    i = len(tokens)
    while i > 1 and is_upper_token(tokens[i - 1]):
        i -= 1
    if i == len(tokens):
        # Aucun token majuscule trouvé : on suppose que le dernier mot est le nom.
        return ' '.join(tokens[:-1]), tokens[-1]
    return ' '.join(tokens[:i]), ' '.join(tokens[i:])


def clean_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def to_iso_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    return None


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else r'C:\Users\SiradioDIALLO\Downloads\RH - CCG - KPI Excel_v2.xlsx'
    out = sys.argv[2] if len(sys.argv) > 2 else 'employees_import.json'

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb['Saisie_Employes']

    rows = []
    skipped_no_name = 0
    for r in range(4, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 17)]
        (employee_id, matricule, nom_complet, business_unit, departement, poste, site,
         manager, date_embauche, type_contrat, statut, salaire, telephone, email,
         anciennete, flag) = vals

        nom_complet = clean_str(nom_complet)
        if not nom_complet:
            continue

        matricule = clean_str(matricule)
        if not matricule:
            skipped_no_name += 1
            continue

        prenom, nom = split_nom_complet(nom_complet)

        statut_clean = clean_str(statut) or 'Actif'
        statut_map = {'actif': 'actif', 'inactif': 'inactif', 'sorti': 'sorti'}
        statut_norm = statut_map.get(statut_clean.lower(), 'actif')

        type_contrat_clean = clean_str(type_contrat)

        rows.append({
            'matricule': matricule,
            'nom': nom,
            'prenom': prenom,
            'nom_complet_source': nom_complet,
            'business_unit_raw': clean_str(business_unit),
            'departement': clean_str(departement),
            'poste': clean_str(poste),
            'site': clean_str(site),
            'manager': clean_str(manager),
            'date_embauche': to_iso_date(date_embauche),
            'type_contrat': type_contrat_clean,
            'statut': statut_norm,
            'salaire_mensuel': salaire if isinstance(salaire, (int, float)) else None,
            'telephone': clean_str(telephone),
            'email': clean_str(email),
        })

    with open(out, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    print(f'{len(rows)} employés extraits -> {out} (lignes sans matricule ignorées : {skipped_no_name})')

    # Aperçu des valeurs distinctes utiles pour valider le mapping BU/site/contrat côté import.
    for key in ('business_unit_raw', 'site', 'type_contrat', 'statut'):
        vals = sorted(set(r[key] for r in rows if r[key]))
        print(f'{key}: {vals}')


if __name__ == '__main__':
    main()
